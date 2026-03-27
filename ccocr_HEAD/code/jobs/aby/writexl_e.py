#!/usr/bin/env python3
# vim: set ts=4 sw=4 sts=4 et ff=unix fenc=utf-8 ai :
#
#   writexl_e.py    230429  cy
#   updated: 260320.182834 by cy
#   updated: 260321 convert webp to jpg in memory for Excel
#   updated: 260321 show engine suffix in pdf cell
#
#--------1---------2---------3---------4---------5---------6---------7--------#

import io
import os
import openpyxl
from openpyxl.utils.cell        import column_index_from_string
from openpyxl.styles.alignment  import Alignment
from openpyxl.styles            import PatternFill
from PIL                        import Image

from m.prnt import prnt
from m.env  import D
from jobs.env import DD

def c2n(c):
    return column_index_from_string(c)

def spic_for_xl(path):
    # Excel does not support webp -- convert to jpg in memory
    if not path.endswith('.webp'):
        return path
    buf = io.BytesIO()
    Image.open(path).convert('RGB').save(buf, format='JPEG')
    buf.seek(0)
    return buf

def writexl_e(dig):
    prnt('writing output excel')
    if not any(len(v) > 0 for v in dig.values()):
        prnt('writexl_e: dig is empty (all OOS), skipping')
        return

    prnt(f'appname: {D.appname}')
    if D.appname == '為替先物レート表_ccocr':
        e_tmplt = os.path.join(
                        os.path.dirname(__file__),'tmplts','e_template.xlsx')
    else:
        e_tmplt = os.path.join(
                        os.path.dirname(__file__),'tmplts','e_template.xlsm')
    prnt(f'e_tmplt\n  {e_tmplt}')

    jobdir  = D.logd
    jobid   = D.jobid
    if D.appname == '為替先物レート表_ccocr':
        wb      = openpyxl.load_workbook(e_tmplt)
    else:
        wb      = openpyxl.load_workbook(e_tmplt, keep_vba = True)
    for docname in dig:
        if len(dig[docname]) == 0:
            continue
        ws = wb.create_sheet()
        ws.title = docname
        ##
        ## sheet header
        ##
        do = dig[docname][0]
        ws.cell(1,1).value = 'doc#'
        ws.cell(1,1).alignment = Alignment(horizontal = 'center')
        ws.cell(1,2).value = 'pdf'
        ws.cell(1,2).alignment = Alignment(horizontal = 'center')
        ws.cell(1,3).value = 'page'
        ws.cell(1,3).alignment = Alignment(horizontal = 'center')
        ws.cell(1,4).value = 'i#'
        ws.cell(1,4).alignment = Alignment(horizontal = 'center')
        gkidx = []
        for cnt,io in enumerate(do.itm):
            ## gkidx (index in 'io' list which has 'gk_')
            if type(io.dl.rg) == str and io.dl.rg.startswith('gk_'):
                gkidx.append(cnt)
            if io.dl.clm == None:
                continue
            ws.cell(1,c2n(io.dl.clm)).value = io.dl.dname
            ws.cell(1,c2n(io.dl.clm)).alignment = Alignment(
                                                        horizontal = 'center')
        ##
        ## sheet body
        ##
        _last_engine = None
        _rie = 0  # row index within engine group
        for row,do in enumerate(dig[docname]):  # do == docObj == 1 biz doc
            if do.engine != _last_engine:
                _rie = 0
                _last_engine = do.engine
            #
            #   fill pdf, document number, page(from), i# (instance num)
            #
            ws.cell(_rie*2+3,1).value = do.docnum
            ws.cell(_rie*2+3,2).value = f'{do.pdf} {do.engine}' if do.engine else do.pdf
            ws.cell(_rie*2+3,3).value = do.fm
            ws.cell(_rie*2+3,4).value = do.inum
            #
            #   fill ocr results
            #
            for clmidx,io in enumerate(do.itm):
                if io.dl.clm == None:
                    continue
                #
                #   ref to papa
                #
                if io.isclone:
                    ws.cell(_rie*2+3,c2n(io.dl.clm)).value = (
                    f'=IF({io.dl.clm}{_rie*2+1}="","",{io.dl.clm}{_rie*2+1})' )
                #
                #   txt and pic
                #
                else:
                    if io.txt == None:
                        txt = ''
                    else:
                        txt = io.txt
                    ws.add_image(   openpyxl.drawing.image.Image(spic_for_xl(io.spic)),
                                    f'{io.dl.clm}{_rie*2+2}')
                    # added 230615
                    ws.cell(
                        _rie*2+2,c2n(io.dl.clm)).value = f'{io.page}_{io.node}'
                    ws.cell(_rie*2+3,c2n(io.dl.clm)).value = txt
                    ws.cell(_rie*2+3,c2n(io.dl.clm)).fill = PatternFill(
                                        patternType='solid', fgColor='dcdcdc')
            _rie += 1
    if len(wb.worksheets) > 1:
        wb.remove(wb['Sheet'])
#    if tmplt != None:
#        return wb
    if D.appname == '為替先物レート表_ccocr':
        wb.save(os.path.join(jobdir,f'{jobid}.xlsx'))
    elif DD.use_web:
        # both mode: distinguish from web output
        wb.save(os.path.join(jobdir,f'{jobid}_xl.xlsm'))
    else:
        wb.save(os.path.join(jobdir,f'{jobid}.xlsm'))
    return
