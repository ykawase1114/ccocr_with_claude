#!/usr/bin/env python3
# vim: set ts=4 sw=4 sts=4 et ff=unix fenc=utf-8 ai :
#
#   writexl_dbg.py
#   updated: 260321 show engine in pdf column
#
#--------1---------2---------3---------4---------5---------6---------7--------#

import os
import openpyxl
from openpyxl.utils.cell        import get_column_letter
from openpyxl.styles.alignment  import Alignment
from openpyxl.styles            import PatternFill
from pprint                     import pformat as pf

from m.prnt import prnt
from m.env  import D

def n2c(n):
    return get_column_letter(n)

def writexl_dbg(dig):
    prnt('writing debug excel')
    if not any(len(v) > 0 for v in dig.values()):
        prnt('writexl_dbg: dig is empty (all OOS), skipping')
        return
    jobdir      = D.logd
    jobid       = D.jobid
    dbg_tmplt   = os.path.join(
                    os.path.dirname(__file__),'tmplts','dbg_template.xlsm')
    wb = openpyxl.load_workbook(dbg_tmplt, keep_vba = True)
    for docname in dig:
        if len(dig[docname]) == 0:
            continue
        ws = wb.create_sheet()
        ws.title = docname
        ##
        ## sheet header
        ##
        do = dig[docname][0]
        # fixed
        ws.cell(1,1).value = 'doc#'
        ws.cell(1,1).alignment = Alignment(horizontal = 'center')
        ws.cell(1,2).value = 'pdf'
        ws.cell(1,2).alignment = Alignment(horizontal = 'center')
        ws.cell(1,3).value = 'page'
        ws.cell(1,3).alignment = Alignment(horizontal = 'center')
        ws.cell(1,4).value = 'inum'
        ws.cell(1,4).alignment = Alignment(horizontal = 'center')
        clm = 4
        for io in do.itm:
            # sqltxt
            clm += 1
            if io.isclone == False:
                ws.cell(1,clm).value = f'{io.dl.dname} sqltxt'
                ws.cell(1,clm).alignment = Alignment(horizontal = 'center')
            # sqlarg
            clm += 1
            ws.cell(1,clm).value = f'{io.dl.dname} sqlarg'
            ws.cell(1,clm).alignment = Alignment(horizontal = 'center')
            # sqlrtn
            clm += 1
            ws.cell(1,clm).value = f'{io.dl.dname} sqlrtn'
            ws.cell(1,clm).alignment = Alignment(horizontal = 'center')
            # regrtn
            clm += 1
            ws.cell(1,clm).value = f'{io.dl.dname} regrtn'
            ws.cell(1,clm).alignment = Alignment(horizontal = 'center')
            # posrtn
            clm += 1
            ws.cell(1,clm).value = f'{io.dl.dname} posrtn'
            ws.cell(1,clm).alignment = Alignment(horizontal = 'center')
            # resubrtn
            clm += 1
            ws.cell(1,clm).value = f'{io.dl.dname} resubrtn'
            ws.cell(1,clm).alignment = Alignment(horizontal = 'center')
        ##
        ## sheet body
        ##
        for row,do in enumerate(dig[docname]):
            # fixed
            ws.cell(row+2,1).value = do.docnum
            ws.cell(row+2,2).value = f'{do.pdf} {do.engine}' if do.engine else do.pdf
            ws.cell(row+2,3).value = do.fm  # page
            ws.cell(row+2,4).value = do.inum
            clm = 4
            for io in do.itm:
                # sqltxt
                clm += 1
                if not io.isclone:
                    ws.cell(row+2,clm).value = io.sqltxt
                    ws.cell(row+2,clm).alignment = Alignment(wrapText = True)
                    ws.column_dimensions[n2c(clm)].width = 255 # max
                else:
                    ws.cell(row+2,clm).fill = PatternFill(
                                        patternType='solid', fgColor='dcdcdc')
                # sqlarg
                clm += 1
                if not io.isclone:
                    ws.cell(row+2,clm).value = pf(io.sqlarg)
                    ws.cell(row+2,clm).alignment = Alignment(wrapText = True)
                    ws.column_dimensions[n2c(clm)].width = 255 # max
                else:
                    ws.cell(row+2,clm).fill = PatternFill(
                                        patternType='solid', fgColor='dcdcdc')
                # sqlrtn
                clm += 1
                if not io.isclone:
                    io.sqlrtn = '\n'.join(
                            [ ','.join([str(ii) for ii in list(i)])
                            for i in io.sqlrtn])
                    ws.cell(row+2,clm).value = io.sqlrtn
                    ws.cell(row+2,clm).alignment = Alignment(wrapText = True)
                    ws.column_dimensions[n2c(clm)].width = 255 # max
                else:
                    ws.cell(row+2,clm).fill = PatternFill(
                                        patternType='solid', fgColor='dcdcdc')
                # regrtn
                clm += 1
                if not io.isclone:
                    io.regrtn = '\n'.join(
                            [ ','.join([str(ii) for ii in list(i)])
                            for i in io.regrtn])
                    ws.cell(row+2,clm).value = io.regrtn
                    ws.cell(row+2,clm).alignment = Alignment(wrapText = True)
                    ws.column_dimensions[n2c(clm)].width = 255 # max
                else:
                    ws.cell(row+2,clm).fill = PatternFill(
                                        patternType='solid', fgColor='dcdcdc')
                # posrtn
                clm += 1
                if not io.isclone:
                    if type(io.posrtn) == list:
                        io.posrtn = ','.join([str(i) for i in io.posrtn])
                    ws.cell(row+2,clm).value = io.posrtn
                    ws.cell(row+2,clm).alignment = Alignment(wrapText = True)
                    ws.column_dimensions[n2c(clm)].width = 255 # max
                else:
                    ws.cell(row+2,clm).fill = PatternFill(
                                        patternType='solid', fgColor='dcdcdc')
                # resubrtn
                clm += 1
                if not io.isclone:
                    if type(io.resubrtn) == list:
                        io.resubrtn = ','.join([str(i) for i in io.resubrtn])
                    ws.cell(row+2,clm).value = io.resubrtn
                    ws.cell(row+2,clm).alignment = Alignment(wrapText = True)
                    ws.column_dimensions[n2c(clm)].width = 255 # max
                else:
                    ws.cell(row+2,clm).fill = PatternFill(
                                        patternType='solid', fgColor='dcdcdc')
    if len(wb.worksheets) > 1:
        wb.remove(wb['Sheet'])
    wb.save(os.path.join(jobdir,f'{jobid}.debug.xlsm'))
    return
