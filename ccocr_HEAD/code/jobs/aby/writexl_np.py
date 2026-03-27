#!/usr/bin/env python3
# vim: set ts=4 sw=4 sts=4 et ff=unix fenc=utf-8 ai :
#
#   writexl_np.py       250806  cy
#   updated: 260320.151037 by cy
#   updated: 260321 show engine suffix in pdf cell
#
#--------1---------2---------3---------4---------5---------6---------7--------#

import os
import openpyxl
from openpyxl.utils.cell        import column_index_from_string
from openpyxl.styles.alignment  import Alignment
from openpyxl.styles            import PatternFill

from m.prnt import prnt
from m.env  import D

def c2n(c):
    return column_index_from_string(c)

def writexl_np(dig):
    prnt('writing output excel')
    if not any(len(v) > 0 for v in dig.values()):
        prnt('writexl_np: dig is empty (all OOS), skipping')
        return
    tmplt   = os.path.join(
                os.path.dirname(__file__), 'tmplts', 'np_template.xlsm')
    jobdir  = D.logd
    jobid   = D.jobid
    wb      = openpyxl.load_workbook(tmplt, keep_vba=True)

    for docname in dig:
        if len(dig[docname]) == 0:
            continue
        ws = wb.create_sheet()
        ws.title = docname
        ##
        ## sheet header
        ##
        do = dig[docname][0]
        ws.cell(1,1).value = 'doc#'
        ws.cell(1,1).alignment = Alignment(horizontal = 'center')
        ws.cell(1,2).value = 'pdf'
        ws.cell(1,2).alignment = Alignment(horizontal = 'center')
        ws.cell(1,3).value = 'page'
        ws.cell(1,3).alignment = Alignment(horizontal = 'center')
        ws.cell(1,4).value = 'i#'
        ws.cell(1,4).alignment = Alignment(horizontal = 'center')
        gkidx = []
        for cnt,io in enumerate(do.itm):
            ## gkidx (index in 'io' list which has 'gk_')
            if type(io.dl.rg) == str and io.dl.rg.startswith('gk_'):
                gkidx.append(cnt)
            if io.dl.clm == None:
                continue
            ws.cell(1,c2n(io.dl.clm)).value = io.dl.dname
            ws.cell(1,c2n(io.dl.clm)).alignment = Alignment(
                                                        horizontal = 'center')
        ##
        ## sheet body
        ##
        for row,do in enumerate(dig[docname]):  # do == docObj == 1 biz doc
            #
            #   fill pdf, document number, page(from), i# (instance num)
            #
            ws.cell(row+2,1).value = do.docnum
            ws.cell(row+2,2).value = f'{do.pdf} {do.engine}' if do.engine else do.pdf
            ws.cell(row+2,3).value = do.fm
            ws.cell(row+2,4).value = do.inum
            #
            #   fill ocr results
            #
            for clmidx,io in enumerate(do.itm):
                if io.dl.clm == None:
                    continue
                #
                #   ref to papa
                #

#                if io.isclone:
#                    ws.cell(row+2,c2n(io.dl.clm)).value = (
#                    f'=IF({io.dl.clm}{row*2+1}="","",{io.dl.clm}{row*2+1})' )

                # updated: 260320.145132 by cy (claud.ai)
                if io.isclone:
                    ws.cell(row+2, c2n(io.dl.clm)).value = (
                        f'=IF({io.dl.clm}{row+1}="","",{io.dl.clm}{row+1})')

                #
                #   txt and pic
                #
                else:
                    if io.txt == None:
                        txt = ''
                    else:
                        txt = io.txt
                    ws.cell(row+2,c2n(io.dl.clm)).value = txt
    if len(wb.worksheets) > 1:
        wb.remove(wb['Sheet'])
    wb.save(os.path.join(jobdir,f'{jobid}.xlsm'))

    return
