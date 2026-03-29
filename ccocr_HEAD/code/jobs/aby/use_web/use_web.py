#!/usr/bin/env python3
# vim: set ts=4 sw=4 sts=4 et ff=unix fenc=utf-8 ai :
#
#   use_web.py      251231  cy
#   updated: 260320.165819 by cy
#   updated: 260321 write header row from htmlsrc.json
#   updated: 260321 text-only rows (no image rows in web output excel)
#
#--------1---------2---------3---------4---------5---------6---------7--------#

import json
import os
import subprocess
import tempfile
import time
from urllib.request import urlopen

import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options

from m.prnt     import prnt
from m.env      import D
from jobs.env   import DD

def use_web():
    websession()
    with open(os.path.join(DD.flskrt,'checked.json'),encoding='utf-8') as f:
        jsn = json.load(f)
    if len(jsn) == 0:
        raise Exception((   '読取結果がブランクです。\n'
                            '読取設定エクセルを直さないと'
                            'いけないかもです。'    ))
    # load header from htmlsrc.json saved by btns.py
    htmlsrc_path = os.path.join(D.logd, 'htmlsrc.json')
    with open(htmlsrc_path, encoding='utf-8') as f:
        htmlsrc = json.load(f)
    tmplt = os.path.normpath(os.path.join(
            os.path.dirname(__file__), r'..', 'tmplts', 'np_template.xlsm'))
    wb = openpyxl.load_workbook(tmplt, keep_vba=True)
    for sht in jsn:
        ws = wb.create_sheet(title=sht)
        # write header row
        hdr = htmlsrc.get(sht, [{}])[0].get('hdr', [])
        hdr = [h if h is not None else '' for h in hdr]
        for ci, h in enumerate(hdr):
            ws.cell(row=1, column=ci+1).value = h
        # write data rows (text only, no image rows)
        for ri, rd in enumerate(jsn[sht]):
            for ci, cd in enumerate(rd):
                ws.cell(row=ri+2, column=ci+1).value = cd
    wb.remove(wb['Sheet'])
    xl = os.path.join(D.logd,f'{D.jobid}_WEB.xlsm')
    wb.save(xl)
    prnt(f'xl saved\n  {xl}')
    # copy instead of second wb.save() to avoid BytesIO closed-file error
    uxl = os.path.join(DD.thisOutd,f'{D.jobid}_WEB.xlsm')
    import shutil as _shutil
    _shutil.copy(xl, uxl)
    prnt(f'user xl saved\n  {uxl}')
    return

def websession():
    flwd    = D.flwd
    appname = D.appname
    #
    #   start flask server
    #
    prnt((  'starting flask web server'
            f'\n  D.jobid {D.jobid}\n  DD.flskrt {DD.flskrt}'))
    subprocess.Popen([  'python'                                            ,
                        os.path.join(os.path.dirname(__file__),'flsk.py')   ,
                        D.jobid                                             ,
                        DD.flskrt                                           ])
    cnt = 0
    while True:
        cnt += 1
        try:
            res = urlopen('http://127.0.0.1:5001/',timeout=1)
        except Exception as e:
            prnt(f'{cnt}) seems flask NOT ready "{e}"')
            time.sleep(1)
            continue
        prnt(f'{cnt}) flask READY (status {res.status})')
        break
    #
    #   open/close web page
    #
    prof    = tempfile.mkdtemp(prefix="selenium_profile_")
    prnt(f'chrome profile at {prof}')
    prnt('using Selenium Manager (auto ChromeDriver)')
    opts    = Options()
    opts.add_argument(f"--user-data-dir={prof}")
    opts.add_argument("--no-first-run")
    opts.add_argument("--no-default-browser-check")
    opts.add_argument("--disable-extensions")
    opts.add_argument("--new-window")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    # Selenium Manager resolves and auto-updates ChromeDriver from
    # https://googlechromelabs.github.io/chrome-for-testing/
    drv = webdriver.Chrome(options=opts)
    prnt('opening URL')
    drv.get('http://127.0.0.1:5001')
    prnt('keep watching user oper')
    cnt = 0
    while True:
        cnt += 1
        try:
            res = urlopen('http://127.0.0.1:5001',timeout=1)
        except Exception as e:
            prnt(f'{cnt}) seems flask TERMINATED "{e}"')
            break
        prnt(f'{cnt}) flask still alive (status {res.status})')
        time.sleep(1)
    prnt('quitting selenimu driver')
    drv.quit()
    return

