#!/usr/bin/env python3
# vim: set ts=4 sw=4 sts=4 et ff=unix fenc=utf-8 ai :
#
#   ldconfig.py     220612  cy
#   ToDo            nks_nonaka  エラーメール _dd
#
#   updated: 260319.162935 by cy
#
#--------1---------2---------3---------4---------5---------6---------7--------#

import re
import sys
from openpyxl import load_workbook

from m.prnt     import prnt
from jobs.env   import DD
from .load_sht  import load_sht
from .init_chk  import init_chk
from .rebuild.rebuild    import rebuild
from .ldresub.ldresub    import ldresub

from .pchk       import pchk

class resub:
    def __init__( self,dname,ptn,rpl,cnt,flg):
        self.dname  = dname
        self.ptn    = ptn
        self.rpl    = rpl
        self.cnt    = cnt
        self.flg    = flg

def ldconfig(sorter):
    msconfig = DD.mscnf
    ##
    ## (1) docdef = ['sheet name to load',...]
    ##
    ## only pick _dd sheet defined in sorter
    ##
    wb = load_workbook(msconfig)
    docdef = set()
    for sn in wb.sheetnames:
        m = re.match(r'^([^_].*)_dd$',sn)
        if m == None:
            continue
        docdef.add(m.groups()[0])
    sorter_k = set(sorter.keys())
    not_used = docdef - sorter_k
    if len(not_used) > 0:
        prnt(f'{not_used} not in sorter, skip reading')
        docdef = docdef - not_used
    not_enough = sorter_k - docdef
    if len(not_enough) > 0:
        raise Exception(f'ERROR: need _dd for {not_enough}')
    docdef = list(docdef)
    ##
    ## (2) load each sheet
    ##
    ## docdef   = { 'docname' : [ def_line, ... ], ... }
    ## def_line = [ def_item, ... ]
    ##
    docdef = load_sht(msconfig,docdef) ## list -> dict
    init_chk(docdef,sorter)
    ## a) pickup defitms
    defitms = {}
    for docname in docdef:
        defitms[docname] = len(docdef[docname])
    ##
    ## (3) rebuild docdef
    ##
    ## docdef['docname'] = obj
    ##
    ## obj.defitms          == ttl count of dlobj (set at blw 'b)')
    ## obj.defs             == [ dlobj, ... ]
    ## obj.child            == [ dlobj of 1level below, ... ]
    ## obj.child...child    == (deeper level)
    ## obj.child...child    == None  ## No More deeper level
    ##
    ##  dlobj.dname == (name of definition)
    ##  dlobj.dtyp  == (line/word)
    ##  dlobj.val   == (regex)
    ##  dlobj.tgt   == (int)
    ##  dlobj.attr  == (None/hd/gk_/gm_)
    ##  dlobj.op    == (None/op)
    ##  dlobj.clm   == (D/E...)
    ##  dlobj.sp_blw
    ##  dlobj.sp_abv
    ##  ....
    rebuild(docdef)
    ## (4) pandas check
    pchk(docdef)
    ## (5) load resub
    ldresub(wb,docdef)
    ## b) set obj.defitms
    for docname in docdef:
        docdef[docname].defitms = defitms[docname]
    wb.close()
    prnt('_dd _resub parsed')
    return docdef
