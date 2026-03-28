#!/usr/bin/env python3
# vim: set ts=4 sw=4 sts=4 et ff=unix fenc=utf-8 ai :
#
#   loadxl.py   251228  cy
#   updated: 260321.114609 by cy
#
#--------1---------2---------3---------4---------5---------6---------7--------#

from m.env      import D
from m.prnt     import prnt
from m.toys     import askfile

import os

from openpyxl import load_workbook

from .extraitms import extraitms
from m.env      import D
from m.prnt     import prnt
from ..env      import DD
from ..util.msg import noxl

def loadxl():
    if DD.config is None:   # if EMBEDDED set by setupPlus()
        config= askfile(
            'OCR設定ファイル（エクセル）を選んでください。'         ,
            'config.txt'                                            )
        prnt(f'config\n  {config}')
    else:
        config  = DD.config

#   not using control file bat STDIN
#    if D.papaidx is not None:
#        control = os.path.join(D.flwd,'flowcontrol.txt')
#        with open(control, 'w', encoding='utf-8') as f:
#            f.write(f'{D.jobid} subprocess {D.papaidx} loaded confXL')

    DD.config   = config
    D.fpath     = config
    locdir      = os.path.dirname(config)

    try:
        wb = load_workbook(filename=config,read_only=True)
    except Exception as e:
        prnt(f'failed to open ocrconfig\n  {config}\n  {e}\n  quitting')
        noxl(e)
        quit()

    names = wb.defined_names
    try:
        [sht,addr] = names['version'].attr_text.split('!')
    except KeyError:
        raise Exception('OCR設定エクセルのバージョンが古いです')

    xlver = wb[sht][addr].value
    [sht,addr] = names['typ'].attr_text.split('!')
    typ = wb[sht][addr].value
    [sht,addr] = names['dpi'].attr_text.split('!')
    dpi = wb[sht][addr].value
    if dpi == 'default':
        dpi = '200dpi'
    [sht,addr] = names['picqlty'].attr_text.split('!')
    picqlty = wb[sht][addr].value
    [sht,addr] = names['twoup'].attr_text.split('!')
    twoup = wb[sht][addr].value
    [sht,addr] = names['hv'].attr_text.split('!')
    hv = wb[sht][addr].value

    if xlver in ['ccocr_v260317']:  ## extra options to support doc. intelli
        extraitms(wb,names)
    engines     = DD.engines
    pdf2api     = DD.pdf2api
    png2api     = DD.png2api
    use_web     = DD.use_web
    use_macro   = DD.use_macro
    use_spic    = DD.use_spic

    prnt(f'''ocrconf options
  config    {os.path.basename(config)}
  xlver     {xlver}
  typ       {typ}
  dpi       {dpi}
  picqlty   {picqlty}
  twoup     {twoup}
  hv        {hv}
  engines   {engines}
  pdf2api   {pdf2api}
  png2api   {png2api}
  use_web   {use_web}
  use_macro {use_macro}
  use_spic  {use_spic}''')

    typdic = {  '文字起こし'            : 'txt' ,
                '帳票読取'              : 'frm' }
#                '帳票読取設定テスト'    : 'cnf' } # obsolete
    jobtyp = typdic[typ]
    prnt(f'jobtyp {jobtyp}')

    DD.frmopt = {   'dpi'       : dpi       ,
                    'qlty'      : picqlty   ,
                    'engines'   : engines   }

    # images info will be added

    if twoup  == 'NOT 2UP':
        pdf2up = False
    elif twoup == '２ＵＰ':
        pdf2up = True
    else:
        raise Exception(f'unexpected sorter B6 {ws["B6"].value}')

    if hv == '左右並び':
        horiz = True
    elif hv  == '上下並び':
        horiz = False
    else:
        raise Exception(f'unexpected sorter B9 {ws["B9"].value}')
    DD.jobtyp   = jobtyp
    DD.pdf2up   = pdf2up
    DD.horiz    = horiz
    wb.close()
    prnt('config xl closed')
    return

