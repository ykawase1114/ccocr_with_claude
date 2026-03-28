#!/usr/bin/env python3
# vim: set ts=4 sw=4 sts=4 et ff=unix fenc=utf-8 ai :
#
#   dump2db_xl.py   260328  cy
#
#   Write {jobid}.dump.xlsx from {jobid}.dump.db.
#   Called from dump2db() after SQLite write completes.
#
#--------1---------2---------3---------4---------5---------6---------7--------#

import os
import sqlite3

import openpyxl

from m.prnt     import prnt
from m.env      import D
from jobs.env   import DD


def dump2db_xl():
    prnt('writing dump excel')
    dbf  = DD.dbf
    xlf  = os.path.join(D.logd, f'{D.jobid}.dump.xlsx')
    DD.dumpf = xlf
    xlbn = os.path.basename(xlf)

    idx = ['seq', 'pdf', 'engine', 'apisrc',
           'page', 'node', 'typ',
           'pg_top', 'pg_btm',
           'top', 'btm', 'lft', 'ryt', 'txt',
           'note1', 'note2']

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet()
    ws.title = 'dxocr element list'
    ws.append(idx)

    con   = sqlite3.connect(dbf)
    cur   = con.cursor()
    lines = cur.execute('SELECT count(*) FROM elm').fetchone()[0]

    for cnty, row in enumerate(cur.execute('''
        SELECT seq,
               pdf, engine, apisrc,
               page, node, typ, pg_top, pg_btm,
               top, btm, lft, ryt, txt, note1, note2
        FROM elm
        ORDER BY seq''')):
        ws.append(row)
        if cnty % 10000 == 0 and cnty > 0:
            prnt(f'{xlbn} {cnty}/{lines} lines')

    con.close()
    wb.save(xlf)

    if lines > 200000:
        prnt(f'{xlbn} NO AUTO FILTER (too many lines)')
        return

    wb = openpyxl.load_workbook(xlf)
    ws = wb.active
    ws.title = 'dxocr element list'
    ws.freeze_panes = ws['A2']
    ws.auto_filter.ref = 'A:R'
    ws.sheet_view.showGridLines = False
    wb.save(xlf)
