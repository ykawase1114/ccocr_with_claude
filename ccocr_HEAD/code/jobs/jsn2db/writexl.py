#!/usr/bin/env python3
# vim: set ts=4 sw=4 sts=4 et ff=unix fenc=utf-8 ai :
#
#   writexl.py  220711 cy
#   updated: 260321 show engine in pdf column
#   updated: 260322 add usepng column to dump excel
#
#--------1---------2---------3---------4---------5---------6---------7--------#

import os
import sqlite3

import openpyxl

from m.prnt import prnt
from m.env  import D
from ..env  import DD

def writexl():
    prnt(f'wriging dump excel')
    dbf         = DD.dbf
    xlf         = os.path.join(D.logd, f'{D.jobid}.dump.xlsx')
    DD.dumpf    = xlf
    xlbn        = os.path.basename(xlf)
    idx         = [ 'seq', 'pdf', 'page', 'node', 'typ',
                    'pg_top', 'pg_btm',
                    'top', 'btm', 'lft', 'ryt', 'txt',
                    'memo1', 'memo2',
                    'engine', 'usepng' ]
    wb          = openpyxl.Workbook(write_only=True)
    ws          = wb.create_sheet()
    ws.title    = 'dxocr element list'
    ws.append(idx)
    con  = sqlite3.connect(dbf)
    cur  = con.cursor()
    lines = cur.execute('SELECT count(*) FROM elm').fetchone()[0]
    for cnty, row in enumerate(cur.execute('''
        SELECT  seq,
                pdf || CASE WHEN engine != '' THEN ' ' || engine ELSE '' END,
                page, node, typ, pg_top, pg_btm,
                top, btm, lft, ryt, txt, note1, note2,
                engine, usepng
        FROM elm
        ORDER BY seq''')):
        ws.append(row)
        if cnty % 10000 == 0 and cnty > 0:
            prnt(f'{xlbn} {cnty}/{lines} lines')
    wb.save(xlf)
    if lines > 200000:
        prnt(f'{xlbn} NO AUTO FILTER (too many lines)')
        return
    wb = openpyxl.load_workbook(xlf)
    ws = wb.active
    ws.title = 'dxocr element list'
    ws.freeze_panes = ws['A2']
    ws.auto_filter.ref = 'A:R'
    ws.sheet_view.showGridLines = False
    wb.save(xlf)
    return
